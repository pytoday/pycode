#!/usr/bin/env python3
# coding=utf-8
# title          : resource_format.py
# description    :
# author         : JackieTsui
# organization   : pytoday.org
# date           : 2018/12/20 15:26
# email          : jackietsui72@gmail.com
# notes          :
# ==================================================

# Import the module needed to run the script
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

"""
<TABLE DEFINE>
| tb1 | CREATE TABLE `tb1` (
  `id` int(10) unsigned NOT NULL AUTO_INCREMENT COMMENT '无意义',
  `ip` varchar(20) COLLATE utf8mb4_bin NOT NULL DEFAULT '0' COMMENT '机器IP',
  `mod1` varchar(45) COLLATE utf8mb4_bin NOT NULL DEFAULT '运维' COMMENT '一级模块',
  `mod2` varchar(200) COLLATE utf8mb4_bin NOT NULL DEFAULT '0' COMMENT '二级模块',
  `jifang` varchar(200) COLLATE utf8mb4_bin NOT NULL DEFAULT '运维' COMMENT '机房',
  `loads` decimal(6,2) NOT NULL DEFAULT '0.00' COMMENT '负载',
  `cpu` decimal(6,2) NOT NULL DEFAULT '0.00' COMMENT 'CPU使用',
  `mem` decimal(6,2) NOT NULL DEFAULT '0.00' COMMENT '内存',
  `io` decimal(6,2) NOT NULL DEFAULT '0.00' COMMENT 'IO使用率',
  `disk` decimal(6,2) NOT NULL DEFAULT '0.00' COMMENT '磁盘空间',
  PRIMARY KEY (`id`)
) ENGINE=InnoDB AUTO_INCREMENT=5411 DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_bin                                 |
"""


# get mod
def get_mod(cur):
    # sql = "SELECT name,saving FROM trade WHERE account = '%s' "
    # data = ('13512345678',)
    sql = "select distinct mod1 from tb1"
    cur.execute(sql)
    for row in cur.fetchall():
        yield row[0]


# get idc name
def get_idc(cur):
    sql = "select distinct jifang from tb1"
    cur.execute(sql)
    for row in cur.fetchall():
        yield row[0]


# get detail
def get_detail(cur):
    sql = "select ip,mod1,mod1,jifang,loads,cpu,mem,io,disk from tb1"
    cur.execute(sql)
    data = cur.fetchall()
    return data


# save detail to excel
def save_detail(ws, data):
    detail_title = ['IP', '一级模块', '二级模块', '机房', '1分钟负载', 'CPU使用率', '内存使用率', '磁盘io', '磁盘使用率']

    ws.append(detail_title)
    for row in data:
        ws.append(row)

    for col in range(ord("A"), ord("I")+1):
        ws.column_dimensions[chr(col)].width = 12.0
    font = Font(bold=True, size=12)
    for col in range(ord("A"), ord("I")+1):
        cell = chr(col)+'1'
        ws[cell].font = font


# get idc summary
def get_sum(idc, mods, cur):
    sql_idc_mod = "select '%s', count(1), round(sum(loads)/count(1), 3), round(sum(cpu)/count(1), 3), round(sum(mem)/count(1),3),round(sum(io)/count(1), 3), round(sum(disk)/count(1), 3) from db1.tb1 where jifang='%s' and mod1 like '%s'"
    sql_idc = "select '汇总', count(1), round(sum(loads)/count(1), 3), round(sum(cpu)/count(1), 3), round(sum(mem)/count(1),3),round(sum(io)/count(1), 3), round(sum(disk)/count(1), 3) from db1.tb1 where jifang='%s'"
    mod_data = []
    for mod in mods:
        cur.execute(sql_idc_mod % (mod, idc, mod))
        mod_data += cur.fetchall()
    cur.execute(sql_idc % (idc,))
    idc_data = cur.fetchall()
    return mod_data, idc_data


# get all summary
def get_allsum(mods, cur):
    sql_idc_mod = "select '%s', count(1), round(sum(loads)/count(1), 3), round(sum(cpu)/count(1), 3), round(sum(mem)/count(1),3),round(sum(io)/count(1), 3), round(sum(disk)/count(1), 3) from db1.tb1 where mod1 like '%s'"
    sql_idc = "select '汇总', count(1), round(sum(loads)/count(1), 3), round(sum(cpu)/count(1), 3), round(sum(mem)/count(1),3),round(sum(io)/count(1), 3), round(sum(disk)/count(1), 3) from db1.tb1"
    mod_data = []
    for mod in mods:
        cur.execute(sql_idc_mod % (mod, mod))
        mod_data += cur.fetchall()
    cur.execute(sql_idc)
    idc_data = cur.fetchall()
    return mod_data, idc_data


# save summary to excel
def save_allsum(ws, idcs, cur):
    # set summary title
    sum_title = ['业务', '服务器数量', '负载平均值', 'CPU使用率%', '内存使用率%', '磁盘IO平均率%', 'DATA磁盘使用率%']
    empty_line = ['']*7

    # save all idc summary
    mods = get_mod(cursor)
    ws['A1'] = "全部机房"
    ws.append(sum_title)
    allsum_data = get_allsum(mods, cur)
    for row in allsum_data[0]:
        ws.append(row)
    ws.append(allsum_data[1][0])
    ws.append(empty_line)

    # save single idc summary
    for idc in idcs:
        mods = get_mod(cursor)
        idc_name = [idc] + ['']*6
        ws.append(idc_name)
        ws.append(sum_title)
        sum_data = get_sum(idc, mods, cur)
        for row in sum_data[0]:
            ws.append(row)
        ws.append(sum_data[1][0])
        ws.append(empty_line)

    # set style
    for col in range(ord("A"), ord("I")+1):
        ws.column_dimensions[chr(col)].width = 18.0


# set sum style
def style_range(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    bold = Side(border_style="medium", color="000000")
    # double = Side(border_style="double", color="ff0000")
    border = Border(top=bold, left=thin, right=thin, bottom=bold)

    font = Font(bold=True, size=13)
    font1 = Font(bold=True, size=12)

    # set border
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = ws[cell_range]
    # for cell in rows[0]:
    # cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right

    # set idc name style
    idc_name_cell = cell_range.split(":")[0]
    ws[idc_name_cell].font = font
    ws[idc_name_cell].alignment = Alignment(horizontal="center", vertical="center")

    row_prefix = cell_range.split(":")[0][1:]
    f_col = 'A' + str(row_prefix)
    l_col = 'G' + str(row_prefix)
    ws.merge_cells(f_col+':'+l_col)

    title_row_prefix = int(row_prefix)+1
    for col in range(ord("A"), ord("G")+1):
        cell = chr(col)+str(title_row_prefix)
        ws[cell].font = font1


if __name__ == '__main__':
    # db config
    connect = pymysql.connect(
        host='192.168.31.134',
        user='user',
        password='password',
        charset='utf8',
        db='db1'
    )
    cursor = connect.cursor()

    # create workbook
    wb = Workbook()
    dst_file = "20XX年X月份自建机房服务器资源使用情况-V1.0.xlsx"
    ws1 = wb.active
    ws1.title = "汇总信息"
    ws2 = wb.create_sheet(title="汇总详情")

    # add data to work sheet
    save_detail(ws2, get_detail(cursor))
    save_allsum(ws1, get_idc(cursor), cursor)

    # style format
    num_mod = len(list(get_mod(cursor)))
    num_idc = len(list(get_idc(cursor)))

    first = 1
    last = num_mod+3
    increment = num_mod+4
    for idc in range(num_idc+1):
        cell_r = 'A'+str(first)+':'+'G'+str(last)
        style_range(ws1, cell_r)
        first += increment
        last += increment

    # save workbook
    wb.save(filename=dst_file)
